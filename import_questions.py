"""
import_questions.py
===================
Reads 'MCQ FOR DEPLOYMENT V1907.2026.xlsx' and merges corrections from
'Correction MCQ V1208.2026.xlsx', then generates SQL and uploads questions to Supabase.

Categories & Sheets:
  - 'DEPLOYMENT URBAN'   -> tags ['Box Van', 'Urban Delivery']
  - 'DEPLOYMENT CURTAIN' -> tags ['Curtain Side', 'Curtain Sider']
  - 'DEPLOYMENT HAULAGE' -> tags ['Container Haulage']
  - 'DEPLOYMENT CARGO'   -> tags ['General Cargo']

Usage:
    python import_questions.py
"""

import pandas as pd
import openpyxl
import uuid
import json
import os
import re

# ── CONFIG ──────────────────────────────────────────────────────────────────
DEPLOYMENT_EXCEL = r'c:\Users\ACER\SafePass\MCQ FOR DEPLOYMENT V1907.2026.xlsx'
CORRECTION_EXCEL = r'c:\Users\ACER\SafePass\Correction MCQ V1208.2026.xlsx'
OUTPUT_SQL = r'c:\Users\ACER\SafePass\supabase\import_all_1040_questions.sql'

SUPABASE_URL = 'https://qhnnyrpcnlddqoyewwkb.supabase.co'
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '')

SHEET_CONFIG = {
    'DEPLOYMENT URBAN': {
        'category': 'Box Van',
        'driver_categories': ['Box Van', 'Urban Delivery'],
    },
    'DEPLOYMENT CURTAIN': {
        'category': 'Curtain Side',
        'driver_categories': ['Curtain Side', 'Curtain Sider'],
    },
    'DEPLOYMENT HAULAGE': {
        'category': 'Container Haulage',
        'driver_categories': ['Container Haulage'],
    },
    'DEPLOYMENT CARGO': {
        'category': 'General Cargo',
        'driver_categories': ['General Cargo'],
    }
}

REGION = 'MY'
DIFFICULTY = 'intermediate'
ANSWER_MAP = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
# ────────────────────────────────────────────────────────────────────────────

def esc(s: str) -> str:
    """Escape single quotes for SQL."""
    if s is None:
        return ''
    return str(s).replace("'", "''")

def parse_float(val) -> float:
    try:
        v = float(val)
        return 0.0 if pd.isna(v) else v
    except (ValueError, TypeError):
        return 0.0

def parse_int(val, default=1) -> int:
    try:
        v = int(val)
        return v if not pd.isna(float(val)) else default
    except (ValueError, TypeError):
        return default

def load_corrections(corr_path: str) -> dict:
    """Load the 24 corrections from Correction MCQ file."""
    if not os.path.exists(corr_path):
        print(f"⚠️ Correction file not found at: {corr_path}")
        return {}

    wb = openpyxl.load_workbook(corr_path, data_only=True)
    sheet = wb['Sheet1']
    corrections = {}

    for r in range(2, sheet.max_row + 1):
        q_en = sheet.cell(r, 3).value
        if q_en and str(q_en).strip():
            key = str(q_en).strip().lower()
            ca = str(sheet.cell(r, 8).value or '').strip().upper()
            exp = str(sheet.cell(r, 9).value or '').strip()
            exp_ms = str(sheet.cell(r, 18).value or '').strip() if sheet.max_column >= 18 else ''
            
            corrections[key] = {
                'ref': sheet.cell(r, 2).value,
                'ca': ca,
                'exp': exp,
                'exp_ms': exp_ms,
            }

    print(f"[OK] Loaded {len(corrections)} corrections from {os.path.basename(corr_path)}")
    return corrections


def parse_all_questions() -> list:
    corrections = load_corrections(CORRECTION_EXCEL)
    wb = openpyxl.load_workbook(DEPLOYMENT_EXCEL, data_only=True)
    
    all_questions = []

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' not found in {DEPLOYMENT_EXCEL}")
            continue

        sheet = wb[sheet_name]
        headers = [str(sheet.cell(1, c).value or '').strip() for c in range(1, sheet.max_column + 1)]
        has_ref = 'reference number' in headers
        
        batch_col = headers.index('Batch No.(JT)') + 1 if 'Batch No.(JT)' in headers else 1
        
        if has_ref:
            q_en_col = 3
            o1_col, o2_col, o3_col, o4_col = 4, 5, 6, 7
            ca_col = 8
            exp_col = 9
            oe_col, od_col, pc_col = 10, 11, 12
            q_ms_col = 13
            m1_col, m2_col, m3_col, m4_col = 14, 15, 16, 17
            exp_ms_col = 18
            ref_col = 2
        else:
            q_en_col = 2
            o1_col, o2_col, o3_col, o4_col = 3, 4, 5, 6
            ca_col = 7
            exp_col = 8
            oe_col, od_col, pc_col = 9, 10, 11
            q_ms_col = 12
            m1_col, m2_col, m3_col, m4_col = 13, 14, 15, 16
            exp_ms_col = 17
            ref_col = None

        sheet_questions = 0
        corr_count = 0

        for r in range(2, sheet.max_row + 1):
            q_en = sheet.cell(r, q_en_col).value
            if not q_en or not str(q_en).strip():
                continue

            text_en = str(q_en).strip()
            batch = parse_int(sheet.cell(r, batch_col).value, 1)

            opt1 = str(sheet.cell(r, o1_col).value or '').strip()
            opt2 = str(sheet.cell(r, o2_col).value or '').strip()
            opt3 = str(sheet.cell(r, o3_col).value or '').strip()
            opt4 = str(sheet.cell(r, o4_col).value or '').strip()

            ca_letter = str(sheet.cell(r, ca_col).value or '').strip().upper()
            explanation_en = str(sheet.cell(r, exp_col).value or '').strip()

            oe = parse_float(sheet.cell(r, oe_col).value)
            od = parse_float(sheet.cell(r, od_col).value)
            pc = parse_float(sheet.cell(r, pc_col).value)

            text_ms = str(sheet.cell(r, q_ms_col).value or '').strip()
            m1 = str(sheet.cell(r, m1_col).value or '').strip()
            m2 = str(sheet.cell(r, m2_col).value or '').strip()
            m3 = str(sheet.cell(r, m3_col).value or '').strip()
            m4 = str(sheet.cell(r, m4_col).value or '').strip()
            explanation_ms = str(sheet.cell(r, exp_ms_col).value or '').strip() if sheet.max_column >= exp_ms_col else ''

            ref_num = parse_int(sheet.cell(r, ref_col).value, 0) if ref_col else None

            # Clean 'nan' values
            def clean_str(v): return '' if str(v).lower() == 'nan' else str(v)
            text_ms = clean_str(text_ms)
            m1, m2, m3, m4 = clean_str(m1), clean_str(m2), clean_str(m3), clean_str(m4)
            explanation_ms = clean_str(explanation_ms)

            # Check and apply corrections
            q_key = text_en.lower()
            if q_key in corrections:
                corr = corrections[q_key]
                if corr['ca']:
                    ca_letter = corr['ca']
                if corr['exp']:
                    explanation_en = corr['exp']
                if corr['exp_ms']:
                    explanation_ms = corr['exp_ms']
                corr_count += 1

            correct_index = ANSWER_MAP.get(ca_letter, 0)

            q_obj = {
                'id': str(uuid.uuid4()),
                'reference_number': ref_num if (ref_num and ref_num > 0) else None,
                'text': text_en,
                'text_ms': text_ms or None,
                'options': [opt1, opt2, opt3, opt4],
                'options_ms': [m1, m2, m3, m4] if (m1 or m2 or m3 or m4) else None,
                'correct_option_index': correct_index,
                'explanation': explanation_en,
                'explanation_ms': explanation_ms or None,
                'regions': [REGION],
                'category': cfg['category'],
                'driver_categories': cfg['driver_categories'],
                'difficulty': DIFFICULTY,
                'batch_number': batch,
                'component_weights': {
                    'operation': oe,
                    'discipline': od,
                    'professionalism': pc,
                }
            }

            all_questions.append(q_obj)
            sheet_questions += 1

        print(f"  * {sheet_name:22s} -> {sheet_questions} questions parsed (Corrections applied: {corr_count})")

    return all_questions


def generate_sql(all_questions: list[dict]) -> str:
    """Generate clean SQL script to replace all questions with the 1,040 updated records."""
    lines = [
        "-- ==========================================================================",
        "-- SafePass Questions Migration - All 1,040 Questions with Merged Corrections",
        "-- ==========================================================================",
        "-- Generated automatically from 'MCQ FOR DEPLOYMENT V1907.2026.xlsx'",
        "-- and 'Correction MCQ V1208.2026.xlsx'",
        "",
        "-- Ensure required columns exist",
        "ALTER TABLE public.questions ADD COLUMN IF NOT EXISTS driver_categories TEXT[];",
        "ALTER TABLE public.questions ADD COLUMN IF NOT EXISTS reference_number INTEGER;",
        "ALTER TABLE public.questions ADD COLUMN IF NOT EXISTS explanation_ms TEXT;",
        "ALTER TABLE public.questions ADD COLUMN IF NOT EXISTS text_ms TEXT;",
        "ALTER TABLE public.questions ADD COLUMN IF NOT EXISTS options_ms JSONB;",
        "",
        "-- Clear existing questions table and replace with authoritative 1,040 questions",
        "TRUNCATE TABLE public.questions;",
        "",
    ]

    for q in all_questions:
        options_json = json.dumps(q['options']).replace("'", "''")
        options_ms_json = json.dumps(q['options_ms']).replace("'", "''") if q['options_ms'] else 'NULL'
        weights_json = json.dumps(q['component_weights']).replace("'", "''")
        cats_arr = "ARRAY['" + "','".join(q['driver_categories']) + "']"
        regions_arr = "ARRAY['" + "','".join(q['regions']) + "']"

        text_ms_val = f"'{esc(q['text_ms'])}'" if q['text_ms'] else 'NULL'
        expl_ms_val = f"'{esc(q['explanation_ms'])}'" if q['explanation_ms'] else 'NULL'
        opts_ms_val = f"'{options_ms_json}'::jsonb" if q['options_ms'] else 'NULL'
        ref_num_val = q['reference_number'] if q['reference_number'] is not None else 'NULL'

        lines.append(f"""INSERT INTO public.questions (
    id, reference_number, text, text_ms, options, options_ms, correct_option_index,
    explanation, explanation_ms, regions, category, driver_categories,
    difficulty, batch_number, component_weights
) VALUES (
    '{q['id']}',
    {ref_num_val},
    '{esc(q['text'])}',
    {text_ms_val},
    '{options_json}'::jsonb,
    {opts_ms_val},
    {q['correct_option_index']},
    '{esc(q['explanation'])}',
    {expl_ms_val},
    {regions_arr},
    '{esc(q['category'])}',
    {cats_arr},
    '{q['difficulty']}',
    {q['batch_number']},
    '{weights_json}'::jsonb
);""")

    return "\n".join(lines)


def main():
    print("=" * 60)
    print("SafePass 1,040 Questions Importer & SQL Generator")
    print("=" * 60)

    all_questions = parse_all_questions()
    print(f"\nTotal questions parsed across all 4 categories: {len(all_questions)}")

    # Category breakdown
    print("\nBreakdown by Category & Batch:")
    for cat_name, cfg in SHEET_CONFIG.items():
        cat = cfg['category']
        qs = [q for q in all_questions if q['category'] == cat]
        print(f"\n  -> {cat} (Total: {len(qs)}):")
        for b in range(1, 9):
            b_count = sum(1 for q in qs if q['batch_number'] == b)
            print(f"     Batch {b}: {b_count} questions")

    # Write SQL Migration
    print(f"\nWriting SQL to: {OUTPUT_SQL}")
    sql_content = generate_sql(all_questions)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"[OK] Successfully generated {OUTPUT_SQL} ({len(sql_content)} bytes)")

    # Also write to import_questions_generated.sql for consistency
    generated_path = r'c:\Users\ACER\SafePass\supabase\import_questions_generated.sql'
    with open(generated_path, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"[OK] Updated {generated_path}")

    print("\nMigration file ready!")

if __name__ == '__main__':
    main()
