import openpyxl

wb_corr = openpyxl.load_workbook('Correction MCQ V1208.2026.xlsx', data_only=True)
sheet_corr = wb_corr['Sheet1']
corr_map = {}
for r in range(2, sheet_corr.max_row + 1):
    ref = sheet_corr.cell(r, 2).value
    q_text = sheet_corr.cell(r, 3).value
    ca = sheet_corr.cell(r, 8).value
    if q_text:
        corr_map[str(q_text).strip().lower()] = {
            'ref': ref,
            'batch': sheet_corr.cell(r, 1).value,
            'ca': ca,
            'row': r,
        }

wb_dep = openpyxl.load_workbook('MCQ FOR DEPLOYMENT V1907.2026.xlsx', data_only=True)
print("=== COMPARING DEPLOYMENT SHEETS WITH CORRECTION MCQ ===")
for sheet_name in wb_dep.sheetnames:
    sheet = wb_dep[sheet_name]
    diff_list = []
    for r in range(2, sheet.max_row + 1):
        # In Deployment sheets, check column for question text and CA
        # Let's inspect headers of this sheet
        q_col = 3 if sheet.cell(1, 2).value == 'reference number' else 2
        ca_col = q_col + 5
        
        q_text = str(sheet.cell(r, q_col).value or '').strip().lower()
        if q_text in corr_map:
            corr_item = corr_map[q_text]
            dep_ca = str(sheet.cell(r, ca_col).value or '').strip()
            corr_ca = str(corr_item['ca'] or '').strip()
            if dep_ca != corr_ca:
                diff_list.append((corr_item['ref'], dep_ca, corr_ca))
    print(f"Sheet '{sheet_name}': {len(diff_list)} Correct Answer discrepancies with Correction MCQ.")
    for ref_id, d_ca, c_ca in diff_list:
        print(f"   Ref {ref_id}: Deployment CA='{d_ca}' vs Correction CA='{c_ca}'")
