import openpyxl
import re
import json

print("=== VERIFYING PARSER & MERGING CORRECTIONS ===")

# 1. Load Corrections
wb_corr = openpyxl.load_workbook('Correction MCQ V1208.2026.xlsx', data_only=True)
sheet_corr = wb_corr['Sheet1']
corrections = {}
for r in range(2, sheet_corr.max_row + 1):
    q_en = sheet_corr.cell(r, 3).value
    if q_en:
        key = str(q_en).strip().lower()
        ca = str(sheet_corr.cell(r, 8).value or '').strip().upper()
        exp = str(sheet_corr.cell(r, 9).value or '').strip()
        corrections[key] = {
            'ref': sheet_corr.cell(r, 2).value,
            'ca': ca,
            'exp': exp,
            'opt1': sheet_corr.cell(r, 4).value,
            'opt2': sheet_corr.cell(r, 5).value,
            'opt3': sheet_corr.cell(r, 6).value,
            'opt4': sheet_corr.cell(r, 7).value,
            'q_ms': sheet_corr.cell(r, 13).value,
            'opt1_ms': sheet_corr.cell(r, 14).value,
            'opt2_ms': sheet_corr.cell(r, 15).value,
            'opt3_ms': sheet_corr.cell(r, 16).value,
            'opt4_ms': sheet_corr.cell(r, 17).value,
            'exp_ms': sheet_corr.cell(r, 18).value,
        }
print(f"Loaded {len(corrections)} corrections.")

# 2. Parse Deployment File
wb_dep = openpyxl.load_workbook('MCQ FOR DEPLOYMENT V1907.2026.xlsx', data_only=True)
SHEET_MAP = {
    'DEPLOYMENT URBAN': ['Box Van', 'Urban Delivery'],
    'DEPLOYMENT CURTAIN': ['Curtain Side', 'Curtain Sider'],
    'DEPLOYMENT HAULAGE': ['Container Haulage'],
    'DEPLOYMENT CARGO': ['General Cargo']
}

all_parsed = []
answer_map = {'A': 0, 'B': 1, 'C': 2, 'D': 3}

for sheet_name, categories in SHEET_MAP.items():
    sheet = wb_dep[sheet_name]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    print(f"\nProcessing {sheet_name} (max_row={sheet.max_row})...")
    
    # Identify column indexes
    # In DEPLOYMENT CARGO: Batch No.(JT) | reference number | Question Text (English) | Option 1 | Option 2 | Option 3 | Option 4 | CA | Explanation (English) | OE | OD | PC | Malay Question Text | Malay Option 1 | Malay Option 2 | Malay Option 3 | Malay Option 4 | Malay Explanation
    # In DEPLOYMENT CURTAIN / URBAN / HAULAGE: Batch No.(JT) | Question Text (English) | Option 1 | Option 2 | Option 3 | Option 4 | CA | Explanation (English) | OE | OD | PC | Malay Question Text | Malay Option 1 | Malay Option 2 | Malay Option 3 | Malay Option 4 | Malay Explanation
    has_ref = 'reference number' in headers
    batch_idx = headers.index('Batch No.(JT)') + 1 if 'Batch No.(JT)' in headers else 1
    
    if has_ref:
        q_en_idx = 3
        o1_idx, o2_idx, o3_idx, o4_idx = 4, 5, 6, 7
        ca_idx = 8
        exp_idx = 9
        oe_idx, od_idx, pc_idx = 10, 11, 12
        q_ms_idx = 13
        m1_idx, m2_idx, m3_idx, m4_idx = 14, 15, 16, 17
        exp_ms_idx = 18
    else:
        q_en_idx = 2
        o1_idx, o2_idx, o3_idx, o4_idx = 3, 4, 5, 6
        ca_idx = 7
        exp_idx = 8
        oe_idx, od_idx, pc_idx = 9, 10, 11
        q_ms_idx = 12
        m1_idx, m2_idx, m3_idx, m4_idx = 13, 14, 15, 16
        exp_ms_idx = 17

    sheet_count = 0
    corr_applied = 0
    for r in range(2, sheet.max_row + 1):
        q_en = sheet.cell(r, q_en_idx).value
        if not q_en or str(q_en).strip() == '':
            continue
        
        batch = int(sheet.cell(r, batch_idx).value or 1)
        o1 = str(sheet.cell(r, o1_idx).value or '').strip()
        o2 = str(sheet.cell(r, o2_idx).value or '').strip()
        o3 = str(sheet.cell(r, o3_idx).value or '').strip()
        o4 = str(sheet.cell(r, o4_idx).value or '').strip()
        ca = str(sheet.cell(r, ca_idx).value or '').strip().upper()
        exp = str(sheet.cell(r, exp_idx).value or '').strip()
        
        oe = float(sheet.cell(r, oe_idx).value or 0.0) if sheet.cell(r, oe_idx).value else 0.0
        od = float(sheet.cell(r, od_idx).value or 0.0) if sheet.cell(r, od_idx).value else 0.0
        pc = float(sheet.cell(r, pc_idx).value or 0.0) if sheet.cell(r, pc_idx).value else 0.0
        
        q_ms = str(sheet.cell(r, q_ms_idx).value or '').strip()
        m1 = str(sheet.cell(r, m1_idx).value or '').strip()
        m2 = str(sheet.cell(r, m2_idx).value or '').strip()
        m3 = str(sheet.cell(r, m3_idx).value or '').strip()
        m4 = str(sheet.cell(r, m4_idx).value or '').strip()
        exp_ms = str(sheet.cell(r, exp_ms_idx).value or '').strip()
        
        # Check for correction
        q_key = str(q_en).strip().lower()
        if q_key in corrections:
            corr = corrections[q_key]
            if corr['ca']:
                ca = corr['ca']
            if corr['exp']:
                exp = corr['exp']
            if corr['exp_ms']:
                exp_ms = corr['exp_ms']
            corr_applied += 1
            
        correct_idx = answer_map.get(ca, 0)
        
        all_parsed.append({
            'batch': batch,
            'q_en': str(q_en).strip(),
            'options': [o1, o2, o3, o4],
            'correct_index': correct_idx,
            'ca_letter': ca,
            'exp_en': exp,
            'weights': {'operation': oe, 'discipline': od, 'professionalism': pc},
            'q_ms': q_ms,
            'options_ms': [m1, m2, m3, m4],
            'exp_ms': exp_ms,
            'categories': categories,
            'sheet': sheet_name
        })
        sheet_count += 1
        
    print(f"  Parsed {sheet_count} questions from {sheet_name} (Corrections merged: {corr_applied})")

print(f"\nTOTAL PARSED QUESTIONS: {len(all_parsed)}")
