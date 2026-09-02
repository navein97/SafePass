import openpyxl
import re
import json

print("=== 1. ANALYZING EXCEL FILES ===")

def analyze_excel(filename):
    print(f"\n--- File: {filename} ---")
    wb = openpyxl.load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
        
        # Count non-empty questions
        q_count = 0
        batches = {}
        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(r, 1).value or sheet.cell(r, 2).value or sheet.cell(r, 3).value
            if val is not None and str(val).strip() != '':
                q_count += 1
                # Check if batch col exists
                batch_val = None
                if 'Batch No.(JT)' in headers:
                    b_idx = headers.index('Batch No.(JT)') + 1
                    batch_val = sheet.cell(r, b_idx).value
                elif 'Batch' in headers:
                    b_idx = headers.index('Batch') + 1
                    batch_val = sheet.cell(r, b_idx).value
                elif 'batch_number' in headers:
                    b_idx = headers.index('batch_number') + 1
                    batch_val = sheet.cell(r, b_idx).value
                
                if batch_val is not None:
                    batches[batch_val] = batches.get(batch_val, 0) + 1

        print(f"Sheet '{sheet_name}': {q_count} questions. Headers: {headers[:6]}")
        if batches:
            print(f"  Batch distribution: {dict(sorted(batches.items(), key=lambda x: str(x[0])))}")

analyze_excel('MCQ FOR DEPLOYMENT V1907.2026.xlsx')
analyze_excel('Correction MCQ V1208.2026.xlsx')
analyze_excel('quiz_question_template_V1106.2026_NAV_no_yellow_rows.xlsx')

print("\n=== 2. ANALYZING SQL FILES ===")
def analyze_sql(filename):
    print(f"\n--- SQL File: {filename} ---")
    with open(filename, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    # Check count of INSERT INTO questions
    inserts = re.findall(r"INSERT INTO\s+public\.questions|INSERT INTO\s+questions", content, re.IGNORECASE)
    print(f"Total INSERT blocks: {len(inserts)}")
    
    # Count rows in values
    # Each question usually has batch_number and driver_categories
    driver_cat_matches = re.findall(r"ARRAY\[([^\]]*)\]", content)
    cat_counts = {}
    for match in driver_cat_matches:
        cleaned = match.strip().replace("'", "").replace('"', '')
        cat_counts[cleaned] = cat_counts.get(cleaned, 0) + 1
    
    print("Driver Categories count in SQL:")
    for k, v in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"  [{k}]: {v}")

    # Batch numbers
    # Look for batch number occurrences in tuples
    # Let's count by matching regex for values tuple pattern
    # e.g. (..., batch_num, ...)
analyze_sql('supabase/import_all_622_questions.sql')
analyze_sql('supabase/import_questions_generated.sql')
